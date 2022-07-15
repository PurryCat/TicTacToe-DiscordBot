import os.path
import random
import asyncio
from datetime import datetime
import discord
import pandas as pd
from discord.ext import commands
from discord.ext.commands import CommandNotFound
from discord.utils import get
from discord_components import DiscordComponents, Button, ButtonStyle
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime


bot = commands.Bot(command_prefix='')
DiscordComponents(bot)
bot.remove_command('help')

adminid = 259726664236269568
fixedTime = datetime.datetime(2022, 4, 30, 22, 30, 00)
token = ""

@bot.event
async def on_ready():
    print("Bot Online")
    if os.path.exists("tictactoe") == False:
        os.mkdir("tictactoe")
        wb = Workbook()
        wb.save(filename="tictactoe/tictactoe.xlsx")
    elif os.path.exists("tictactoe") == True and os.path.exists("tictactoe/tictactoe.xlsx") == False:
        wb = Workbook()
        wb.save(filename="tictactoe/tictactoe.xlsx")

@bot.listen()
async def on_message(ctx):
    blacklist = [609884159590137927]
    if ctx.author.id in blacklist:
        return
    contentWords = ctx.content.split()
    try:
        contentWords[0] 
    except:
        return

    if ctx.author.bot: 
        return
# ------------------------------------------------------- Help -------------------------------------------------------
    elif contentWords[0].lower() == "fh" or contentWords[0].lower() == "fhelp" or contentWords[0].lower() == "fhelp":
        words = ctx.content.split()
        if len(words) > 1:
            if words[1] == "forfeitttt":
                embedVar = discord.Embed(title="Command Info: ``forfeitttt``", description="Shortcuts: ``ffttt`` \nUsage: ``ffttt`` \n \nForfeits your current game of Tic-Tac-Toe, resulting in a loss", color=0x00ff00)
                
            elif words[1] == "tictactoe":
                embedVar = discord.Embed(title="Command Info: ``tictactoe``", description="Shortcuts: ``ttt``\nUsage: ``fttt`` ``*user*`` \n \nStarts a game of Tic-Tac-Toe with the specified person", color=0x00ff00)
        else:
            embedVar = discord.Embed(title="TicTacToe Help", description="Type ``fhelp command`` to see more details about a particular command.", color=0x00ff00)
            embedVar.add_field(name="Games" , value="``userinfo``, ``tictactoe``, ``forfeitttt``", inline=True)
        await ctx.channel.send(embed=embedVar)
        
# ------------------------------------------------------- User Info -------------------------------------------------------
    if contentWords[0].lower() == "fui" or contentWords[0].lower() == "fuserinfo":
        authorRow = idcheckttt(ctx.author.id)
        tictactoeswon = str(sheet.cell(row=authorRow, column=4).value)
        tictactoeslost = str(sheet.cell(row=authorRow, column=5).value)
        tictactoestied = str(sheet.cell(row=authorRow, column=6).value)
        tictactoesplayed = str(sheet.cell(row=authorRow, column=7).value)

        description = "** Tic-Tac-Toes Won: **" + tictactoeswon + "** \nTic-Tac-Toes Lost: **" + tictactoeslost + "** \nTic-Tac-Toes Tied: **" + tictactoestied + "** \nTic-Tac-Toes Played: **" + tictactoesplayed + "**"
        embedVar = discord.Embed(title="View User Info", description="Showing statistics for " + str(ctx.author.mention) + "\n \n" + description, color=0x00ff00)
        await ctx.channel.send(embed=embedVar)
# ------------------------------------------------------- Resets TicTactoe ------------------------------------------------------
    if contentWords[0].lower() == "frtictactoe" or contentWords[0].lower() == "frttt" and ctx.author.id == adminid:
        words = ctx.content.split()
        words.pop(0)
        if len(words) != 1:
            await ctx.channel.send(str(ctx.author.mention) + " no")
            return
        for user_mentioned in ctx.mentions:
            user = user_mentioned
        if user.id == ctx.author.id:
            await ctx.channel.send(str(ctx.author.mention) + " no")
            return
        id = ctx.author.id
        authorrow = idcheckttt(id)
        userrow = idcheckttt(user.id)
        workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
        sheet = workbook.active
        sheet["B" + str(authorrow)] = 0
        sheet["B" + str(userrow)] = 0
        sheet["C" + str(authorrow)] = None
        sheet["C" + str(userrow)] = None
        workbook.save(filename="tictactoe/tictactoe.xlsx")
        await ctx.channel.send("tictactoe data reset")

# ------------------------------------------------------- Forfeits Tic Tac Toe -------------------------------------------------------
    if contentWords[0].lower() == "ffttt":
        forfeit = [None] * 1000
        ran = random.randrange(0, 1000)
        authorrow = idcheckttt(ctx.author.id)
        workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
        sheet = workbook.active
        alreadystarted = str(sheet.cell(row=authorrow, column=2).value)
        if alreadystarted == 1:
            await ctx.channel.send("You don't have a game of Tic-Tac-Toe currently started")
        else:
            if alreadystarted == "1":
                forfeit[ran] = await ctx.channel.send("Forfeit Tic-Tac-Toe?", components = [[Button(label="Yes", style="3", custom_id="button1" + str(ctx.author.id)), Button(label="No", style="4", custom_id="button2" + str(ctx.author.id))]])
                while True:
                    try:
                        event = await bot.wait_for("button_click", timeout = 30.0)
                    except asyncio.TimeoutError: 
                        if reply != 1:
                            await ctx.channel.send(f"{ctx.author.mention}, you didn't reply fast enough..") 
                            try:
                                await forfeit[ran].delete()
                            except discord.NotFound:
                                return
                            return
                    else:
                        if event.user.id == ctx.author.id:
                            if event.component.id == "button1" + str(ctx.author.id):
                                reply = 1
                                try:
                                    await forfeit[ran].delete()  
                                except discord.NotFound:
                                    return
                                sheet.cell(row=authorrow, column=2).value = 2
                                workbook.save(filename="tictactoe/tictactoe.xlsx")

# ------------------------------------------------------- Main Tic Tac Toe Code -------------------------------------------------------
    if contentWords[0].lower() == "ftictactoe" or contentWords[0].lower() == "fttt":
        words = ctx.content.split()
        words.pop(0)
        if len(words) != 1:
            await ctx.channel.send("Invalid command usage, do ``fhelp`` ``tictactoe`` for more info")
            return
        for user_mentioned in ctx.mentions:
            user = user_mentioned
        if user.id == ctx.author.id:
            await ctx.channel.send(str(ctx.author.mention) + " cant play a game against yourself")
            return
        if user.bot: 
            await ctx.channel.send(str(ctx.author.mention) + " cant play a game against a bot")
            return
        id = ctx.author.id
        authorrow = idcheckttt(id)
        userrow = idcheckttt(user.id)
        ids = int(user.id) + int(ctx.author.id)
        workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
        sheet = workbook.active
        reply = 0
        if str(sheet.cell(row=authorrow, column=2).value) == "1":
            await ctx.channel.send("You already have a game of Tic-Tac-Toe started")
        elif str(sheet.cell(row=userrow, column=2).value) == "1":
            await ctx.channel.send(ctx.author.mention + " That user already has a game of Tic-Tac-Toe started")
        elif str(sheet.cell(row=authorrow, column=2).value) == "2":
            await ctx.channel.send("You already have a challenge of Tic-Tac-Toe pending, cancel the otherone to start a new one")
        else:
            message = [None] * 10000
            ran = random.randrange(0, 10000)
            message[ran]  = await ctx.channel.send(str(user.mention) + " Do you want to play a game of Tic-Tac-Toe vs " + str(ctx.author.mention) + "?", components = [[Button(label="Yes", style="3", custom_id="button1" + str(ctx.author.id)), Button(label="No", style="4", custom_id="button2" + str(ctx.author.id))]])
            sheet["B" + str(authorrow)] = 2
            workbook.save(filename="tictactoe/tictactoe.xlsx")
            while True:
                try:
                    event = await bot.wait_for("button_click", timeout = 30.0)
                except asyncio.TimeoutError: 
                    if reply != 1:
                        await ctx.channel.send(f"{user}, didn't reply fast enough.") 
                        try:
                            await event.respond()
                        except:
                            pass
                        return
                else:
                    if event.component.id == "button2" + str(ctx.author.id):
                        workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
                        sheet = workbook.active
                        sheet["B" + str(authorrow)] = 0
                        workbook.save(filename="tictactoe/tictactoe.xlsx")
                        if event.user.id == user.id or event.user.id == ctx.author.id:
                            if event.user.id == user.id:
                                await ctx.channel.send(content="Tic-Tac-Toe was declined")
                            else:
                                await ctx.channel.send(content="Tic-Tac-Toe was cancelled")
                            try:
                                await message[ran].delete()
                            except discord.NotFound:
                                return
                            return
                    if event.user.id == user.id:
                        if event.component.id == "button1" + str(ctx.author.id):
                            reply = 1
                            try:
                                await message[ran].delete()
                            except discord.NotFound:
                                return
                            workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
                            sheet = workbook.active
                            sheet["B" + str(authorrow)] = 1
                            sheet["B" + str(userrow)] = 1
                            sheet["C" + str(authorrow)] = str(user.id)
                            sheet["C" + str(userrow)] = str(ctx.author.id)
                            workbook.save(filename="tictactoe/tictactoe.xlsx")
                            board = ["."] * 9
                            ran = random.randrange(0, 2)
                            if ran == 0:
                                description = "(X) " + ctx.author.mention + " vs " + user.mention + " **(O)**"
                                players = {
                                    'X': ctx.author.id,
                                    'O': user.id
                                }
                            elif ran == 1:
                                players = {
                                    'O': ctx.author.id,
                                    'X': user.id
                                }
                                description = "**(O)** " + ctx.author.mention + " vs " + user.mention + " (X)"
                            buttons = getbuttons(board, ids)
                            message[ran] = await ctx.channel.send("ribbit", components=buttons)
                            turn = 'X'
                            while True:
                                workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
                                sheet = workbook.active
                                if str(sheet.cell(row=authorrow, column=2).value) == "2":
                                    await ctx.channel.send(str(ctx.author.mention) + " your tic-tac-toe has been forfeited")
                                    await message[ran].edit(str(ctx.author) + " forfeited", components=buttons)
                                    message[ran] = None
                                    workbook.close
                                    workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
                                    sheet = workbook.active
                                    sheet["B" + str(authorrow)] = 0
                                    sheet["B" + str(userrow)] = 0
                                    sheet["C" + str(authorrow)] = None
                                    sheet["C" + str(userrow)] = None
                                    if str(players["X"]) == str(ctx.author.id):
                                        sheet["D" + str(userrow)] = int(sheet.cell(row=userrow, column=4).value) + 1
                                    else:
                                        sheet["D" + str(authorrow)] = int(sheet.cell(row=authorrow, column=5).value) + 1
                                    sheet["G" + str(authorrow)] = int(sheet.cell(row=authorrow, column=7).value) + 1
                                    sheet["G" + str(userrow)] = int(sheet.cell(row=userrow, column=7).value) + 1
                                    workbook.save(filename="tictactoe/tictactoe.xlsx")
                                    break
                                elif str(sheet.cell(row=userrow, column=2).value) == "2":
                                    await ctx.channel.send(str(user.mention) + " your tic-tac-toe has been forfeited")
                                    await message[ran].edit(str(user) + " forfeited", components=buttons)
                                    message[ran] = None
                                    workbook.close
                                    workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
                                    sheet = workbook.active
                                    sheet["B" + str(authorrow)] = 0
                                    sheet["B" + str(userrow)] = 0
                                    sheet["C" + str(authorrow)] = None
                                    sheet["C" + str(userrow)] = None
                                    if str(players["X"]) == str(user.id):
                                        sheet["D" + str(userrow)] = int(sheet.cell(row=userrow, column=4).value) + 1
                                    else:
                                        sheet["D" + str(authorrow)] = int(sheet.cell(row=authorrow, column=5).value) + 1
                                    sheet["G" + str(authorrow)] = int(sheet.cell(row=authorrow, column=7).value) + 1
                                    sheet["G" + str(userrow)] = int(sheet.cell(row=userrow, column=7).value) + 1
                                    workbook.save(filename="tictactoe/tictactoe.xlsx")
                                    break
                                buttons = getbuttons(board, ids)
                                turnid = players[turn]
                                if description == "**(X)** " + ctx.author.mention + " vs " + user.mention + " (O)" or description == "(X) " + ctx.author.mention + " vs " + user.mention + " **(O)**":
                                    if turn == "O":
                                        description = "(X) " + ctx.author.mention + " vs " + user.mention + " **(O)**"
                                    else:
                                        description = "**(X)** " + ctx.author.mention + " vs " + user.mention + " (O)"
                                else:
                                    if turn == "X":
                                        description = "(O) " + ctx.author.mention + " vs " + user.mention + " **(X)**"
                                    else:
                                        description = "**(O)** " + ctx.author.mention + " vs " + user.mention + " (X)"
                                await message[ran].edit(description, components=buttons)
                                event = await bot.wait_for("button_click")
                                try:
                                    await event.respond()
                                except:
                                    pass
                                if str(event.user.id) == str(turnid):
                                    if str(event.component.id[0]) + str(ids) == str(event.component.id):
                                        column = int(event.component.id[0])
                                        if board[column] == ".":
                                            board[column] = str(turn)    
                                            if turn == "X":
                                                turn = "O"
                                            else:
                                                turn = "X"  
                                            win = checkwin(board)
                                            beanss = board.count(".")
                                            if beanss == 0:
                                                buttons = getbuttons(board, ids)
                                                await message[ran].edit(str(description) + "\nTie", components=buttons)
                                                message[ran] = None
                                                workbook.close
                                                workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
                                                sheet = workbook.active
                                                sheet["B" + str(authorrow)] = 0
                                                sheet["B" + str(userrow)] = 0
                                                sheet["C" + str(authorrow)] = None
                                                sheet["C" + str(userrow)] = None
                                                sheet["G" + str(authorrow)] = int(sheet.cell(row=authorrow, column=7).value) + 1
                                                sheet["G" + str(userrow)] = int(sheet.cell(row=userrow, column=7).value) + 1
                                                sheet["F" + str(authorrow)] = int(sheet.cell(row=authorrow, column=6).value) + 1
                                                sheet["F" + str(userrow)] = int(sheet.cell(row=userrow, column=6).value) + 1
                                                workbook.save(filename="tictactoe/tictactoe.xlsx")
                                                break
                                            if win == "X":
                                                buttons = getbuttons(board, ids)
                                                rarity = random.randrange(1, 1000)
                                                if rarity <= 500: #50%
                                                    bug = "**common**"
                                                    column = 3
                                                elif rarity >= 501 and rarity <= 700: #20%
                                                    bug = "**uncommon**"
                                                    column = 4
                                                elif rarity >= 701 and rarity <= 850: #15%
                                                    bug = "**rare**" 
                                                    column = 5
                                                elif rarity >= 851 and rarity <= 950: #10%
                                                    bug = "**epic**" 
                                                    column = 6
                                                elif rarity >= 951 and rarity <= 1000: #5%
                                                    bug = "**legendary**"
                                                    column = 7
                                                await message[ran].edit(str(description) + "\nX has won! and recieved a " + str(bug) + " bug!",  components=buttons)
                                                message[ran] = None
                                                workbook.close
                                                workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
                                                sheet = workbook.active
                                                sheet["B" + str(authorrow)] = 0
                                                sheet["B" + str(userrow)] = 0
                                                sheet["C" + str(authorrow)] = None
                                                sheet["C" + str(userrow)] = None
                                                if str(players["X"]) == str(ctx.author.id):
                                                    sheet["D" + str(authorrow)] = int(sheet.cell(row=authorrow, column=5).value) + 1
                                                    sheet["E" + str(userrow)] = int(sheet.cell(row=userrow, column=4).value) + 1
                                                    winner = ctx.author
                                                else:
                                                    sheet["E" + str(authorrow)] = int(sheet.cell(row=authorrow, column=5).value) + 1
                                                    sheet["D" + str(userrow)] = int(sheet.cell(row=userrow, column=4).value) + 1
                                                    winner = user
                                                sheet["G" + str(authorrow)] = int(sheet.cell(row=authorrow, column=7).value) + 1
                                                sheet["G" + str(userrow)] = int(sheet.cell(row=userrow, column=7).value) + 1
                                                workbook.save(filename="tictactoe/tictactoe.xlsx")
                                                winnerrow = idcheckinv(winner)
                                                workbook = load_workbook(filename="tamagotchi/inventories.xlsx")
                                                sheet = workbook.active
                                                sheet.cell(row=winnerrow, column=column).value = sheet.cell(row=winnerrow, column=column).value + 1
                                                workbook.save(filename="tamagotchi/inventories.xlsx")
                                                break
                                            elif win == "O":
                                                buttons = getbuttons(board, ids)
                                                rarity = random.randrange(1, 1000)
                                                if rarity <= 500: #50%
                                                    bug = "**common**"
                                                    column = 3
                                                elif rarity >= 501 and rarity <= 700: #20%
                                                    bug = "**uncommon**"
                                                    column = 4
                                                elif rarity >= 701 and rarity <= 850: #15%
                                                    bug = "**rare**" 
                                                    column = 5
                                                elif rarity >= 851 and rarity <= 950: #10%
                                                    bug = "**epic**" 
                                                    column = 6
                                                elif rarity >= 951 and rarity <= 1000: #5%
                                                    bug = "**legendary**"
                                                    column = 7
                                                await message[ran].edit(str(description) + "\nO has won! and recieved a " + str(bug) + " bug!",  components=buttons)
                                                message[ran] = None
                                                workbook.close
                                                workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
                                                sheet = workbook.active
                                                sheet["B" + str(authorrow)] = 0
                                                sheet["B" + str(userrow)] = 0
                                                sheet["C" + str(authorrow)] = None
                                                sheet["C" + str(userrow)] = None
                                                if str(players["O"]) == str(ctx.author.id):
                                                    sheet["D" + str(authorrow)] = int(sheet.cell(row=authorrow, column=5).value) + 1
                                                    sheet["E" + str(userrow)] = int(sheet.cell(row=userrow, column=4).value) + 1
                                                    winner = ctx.author
                                                else:
                                                    sheet["E" + str(authorrow)] = int(sheet.cell(row=authorrow, column=5).value) + 1
                                                    sheet["D" + str(userrow)] = int(sheet.cell(row=userrow, column=4).value) + 1
                                                    winner = user
                                                sheet["G" + str(authorrow)] = int(sheet.cell(row=authorrow, column=7).value) + 1
                                                sheet["G" + str(userrow)] = int(sheet.cell(row=userrow, column=7).value) + 1
                                                workbook.save(filename="tictactoe/tictactoe.xlsx")
                                                winnerrow = idcheckinv(winner)
                                                workbook = load_workbook(filename="tamagotchi/inventories.xlsx")
                                                sheet = workbook.active
                                                sheet.cell(row=winnerrow, column=column).value = sheet.cell(row=winnerrow, column=column).value + 1
                                                workbook.save(filename="tamagotchi/inventories.xlsx")
                                                break

# -------------------------------------------------------------------------------------------------------------- Functions --------------------------------------------------------------------------------------------------------------

# ------------------------------------------------------- ID check, input user id, returns row of author in tictactoe.xlsx-------------------------------------------------------
def idcheckttt(id):
    workbook = load_workbook(filename="tictactoe/tictactoe.xlsx")
    sheet = workbook.active
    max_row=sheet.max_row
    found = 0
    for i in range(1,max_row+1):
        cell_obj = sheet.cell(row=i, column=1)
        if cell_obj.value == str(id):
            row = i
            found = 1
        elif i == max_row and cell_obj.value != str(id) and found != 1:
            j = i + 1
            sheet ["A" + str(j)] = str(id)
            sheet ["B" + str(j)] = 0
            sheet ["D" + str(j)] = 0
            sheet ["E" + str(j)] = 0
            sheet ["F" + str(j)] = 0
            sheet ["G" + str(j)] = 0
            workbook.save(filename="tictactoe/tictactoe.xlsx")
            row = j
    workbook.close
    return row

# ------------------------------------------------------- Check board for Tic-Tac-Toe win ------------------------------------------------------
def checkwin(board):
    winner = False
    for x in range(2):
        if x == 1:
            xoro = "X"
        else:
            xoro = "O"
        #Horizontals
        if str(board[0]) == xoro and str(board[1]) == xoro and str(board[2]) == xoro or str(board[3]) == xoro and str(board[4]) == xoro and str(board[5]) == xoro or str(board[6]) == xoro and str(board[7]) == xoro and str(board[8]) == xoro:
            winner = xoro

        #Verticles
        if str(board[0]) == xoro and str(board[3]) == xoro and str(board[6]) == xoro or str(board[1]) == xoro and str(board[4]) == xoro and str(board[7]) == xoro or str(board[2]) == xoro and str(board[5]) == xoro and str(board[8]) == xoro:
            winner = xoro
        
        # Diagonals
        if str(board[2]) == xoro and str(board[4]) == xoro and str(board[6]) == xoro or str(board[0]) == xoro and str(board[4]) == xoro and str(board[8]) == xoro:
            winner = xoro
    return winner

# ------------------------------------------------------- Gets the buttons for Tic Tac Toe ------------------------------------------------------
def getbuttons(board, ids):
    one = Button(style=ButtonStyle.gray, label=str(board[0]), id="0" + str(ids))
    two = Button(style=ButtonStyle.gray, label=str(board[1]), id="1" + str(ids))
    three = Button(style=ButtonStyle.gray, label=str(board[2]), id="2" + str(ids))
    four = Button(style=ButtonStyle.gray, label=str(board[3]), id="3" + str(ids))
    five = Button(style=ButtonStyle.gray, label=str(board[4]), id="4" + str(ids))
    six = Button(style=ButtonStyle.gray, label=str(board[5]), id="5" + str(ids))
    seven = Button(style=ButtonStyle.gray, label=str(board[6]), id="6" + str(ids))
    eight = Button(style=ButtonStyle.gray, label=str(board[7]), id="7" + str(ids))
    nine = Button(style=ButtonStyle.gray, label=str(board[8]), id="8" + str(ids))
    buttons = [[one,two,three],[four,five,six],[seven,eight,nine]]
    return buttons

# ------------------------------------------------------- Gets the bug inventory of the person-------------------------------------------------------
def idcheckinv(ctx):
    if isinstance(ctx, int) == True:
        id = ctx
    else:
        id = ctx.author.id
    workbook = load_workbook(filename="tamagotchi/inventories.xlsx")
    sheet = workbook.active
    max_row=sheet.max_row
    found = 0
    for i in range(1,max_row+1):
        cell_obj = sheet.cell(row=i, column=1)
        if cell_obj.value == str(id):
            row = i
            found = 1
            if isinstance(ctx, int) != True:
                if sheet.cell(row=row, column=2).value != str(ctx.author.name):
                    sheet.cell(row=row, column=2).value = str(ctx.author.name)
        elif i == max_row and cell_obj.value != str(id) and found != 1:
            j = i + 1
            row = j
            sheet ["A" + str(j)] = str(id)
            sheet ["B" + str(j)] = 0
            sheet ["C" + str(j)] = 0
            sheet ["D" + str(j)] = 0
            sheet ["E" + str(j)] = 0
            sheet ["F" + str(j)] = 0
            if isinstance(ctx, int) != True:
                if sheet.cell(row=row, column=2).value != str(ctx.author.name):
                    sheet.cell(row=row, column=2).value = str(ctx.author.name)
        workbook.save(filename="tamagotchi/inventories.xlsx")
    workbook.close
    return row
    
# ------------------------------------------------------- Error checking -------------------------------------------------------
@bot.event
async def on_command_error(ctx, error):
    if isinstance(error, CommandNotFound):
        return

bot.run(token)